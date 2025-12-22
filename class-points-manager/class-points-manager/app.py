from flask import Flask, render_template, jsonify, request, send_file, make_response, session, redirect, url_for
from flask_cors import CORS
from werkzeug.utils import secure_filename
import sqlite3
import json
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import io
import tempfile
import time
import socket
import webbrowser
from config import Config

class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'class-points-manager-secret-2025'
    DATABASE_PATH = os.path.join(os.path.dirname(__file__), 'data', 'class_points.db')
    UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
    DEBUG = True

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = 'cp-manager-secure-key' # 显式设置以便 session 工作
CORS(app)  # 允许跨域请求

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def get_system_password():
    """从本地文件读取系统密码"""
    password_file = os.path.join(os.path.dirname(__file__), 'password.txt')
    if not os.path.exists(password_file):
        with open(password_file, 'w', encoding='utf-8') as f:
            f.write('123456')
        return '123456'
    with open(password_file, 'r', encoding='utf-8') as f:
        return f.read().strip()

@app.before_request
def check_auth():
    """全局登录检查"""
    allowed_paths = ['/login', '/static', '/api/verify_password']
    if any(request.path.startswith(path) for path in allowed_paths):
        return
    if 'logged_in' not in session:
        return redirect(url_for('login_page'))

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """登录页面及处理"""
    if request.method == 'POST':
        data = request.json
        input_pwd = str(data.get('password', '')).strip()
        if input_pwd == get_system_password():
            session['logged_in'] = True
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '密码错误'}), 403
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login_page'))

@app.route('/api/verify_password', methods=['POST'])
def verify_password_api():
    """专门供AJAX调用的验证接口"""
    data = request.json
    input_pwd = str(data.get('password', '')).strip()
    if input_pwd == get_system_password():
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '密码错误'}), 403

# 数据库初始化
def init_db():
    """初始化数据库"""
    try:
        # 确保data目录存在
        data_dir = os.path.join(os.path.dirname(__file__), 'data')
        os.makedirs(data_dir, exist_ok=True)
        print(f"数据库文件将保存在: {app.config['DATABASE_PATH']}")
        
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        cursor = conn.cursor()
        
        # 创建班级表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            teacher TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        
        # 创建分组表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            color TEXT DEFAULT '#667eea',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (class_id) REFERENCES classes (id),
            UNIQUE(class_id, name)
        )
        ''')
        
        # 创建学生表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER,
            group_id INTEGER,
            name TEXT NOT NULL,
            student_id TEXT NOT NULL UNIQUE,
            points INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (class_id) REFERENCES classes (id),
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
        ''')
        
        # 创建积分历史表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS points_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER,
            change_amount INTEGER NOT NULL,
            reason TEXT,
            teacher TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students (id)
        )
        ''')
        
        # 创建分组积分历史表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS group_points_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER,
            change_amount INTEGER NOT NULL,
            reason TEXT,
            teacher TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
        ''')

        # 创建奖品表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS rewards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            points_cost INTEGER NOT NULL,
            image_path TEXT,
            stock INTEGER DEFAULT 10,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # 创建兑换记录表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS redemptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            reward_id INTEGER NOT NULL,
            redeemed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students (id),
            FOREIGN KEY (reward_id) REFERENCES rewards (id)
        )
        ''')

        # 创建小组兑换记录表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS group_redemptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            reward_id INTEGER NOT NULL,
            redeemed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (group_id) REFERENCES groups (id),
            FOREIGN KEY (reward_id) REFERENCES rewards (id)
        )
        ''')

        # 创建家校评价表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS student_evaluations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            content TEXT,
            period TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students (id)
        )
        ''')
        
        conn.commit()
        print("数据库表创建成功")
        
        # 检查表是否真的创建了
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        print(f"数据库中的表: {[t[0] for t in tables]}")
        
        conn.close()
        
    except Exception as e:
        print(f"初始化数据库失败: {e}")
        return False, str(e)
    
    return True, "数据库初始化成功"

def get_db_connection():
    """获取数据库连接"""
    # 先检查数据库文件是否存在
    if not os.path.exists(app.config['DATABASE_PATH']):
        print(f"数据库文件不存在，重新初始化: {app.config['DATABASE_PATH']}")
        init_db()
    
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    conn.row_factory = sqlite3.Row  # 返回字典格式
    return conn

def check_and_init_tables():
    """检查并初始化所有表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查所有必需的表是否存在
        required_tables = ['classes', 'groups', 'students', 'points_history', 'group_points_history', 'group_redemptions']
        existing_tables = []
        
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        existing_tables = [t[0] for t in tables]
        
        print(f"现有表: {existing_tables}")
        print(f"必需表: {required_tables}")
        
        missing_tables = [t for t in required_tables if t not in existing_tables]
        
        if missing_tables:
            print(f"缺少的表: {missing_tables}")
            conn.close()
            # 重新初始化数据库
            success, message = init_db()
            if not success:
                print(f"重新初始化失败: {message}")
                return False
        else:
            print("所有必需表都存在")
        
        conn.close()
        return True
        
    except Exception as e:
        print(f"检查表时出错: {e}")
        return False

# ============ 基础API路由 ============

@app.route('/')
def index():
    """首页"""
    return render_template('index.html')

@app.route('/random')
def random_page():
    """随机点名页面"""
    return render_template('random.html')

@app.route('/report')
def report_page():
    """家校合作/报告页面"""
    return render_template('report.html')

@app.route('/api/students/<int:student_id>/evaluation', methods=['GET', 'POST'])
def handle_evaluation(student_id):
    """获取或保存学生评价"""
    conn = get_db_connection()
    if request.method == 'POST':
        data = request.json
        content = data.get('content')
        period = data.get('period', datetime.now().strftime('%Y-%m'))
        
        conn.execute('''
            INSERT INTO student_evaluations (student_id, content, period)
            VALUES (?, ?, ?)
        ''', (student_id, content, period))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    # GET
    evals = conn.execute('''
        SELECT * FROM student_evaluations 
        WHERE student_id = ? 
        ORDER BY created_at DESC
    ''', (student_id,)).fetchall()
    conn.close()
    return jsonify([dict(e) for e in evals])

@app.route('/api/classes/<int:class_id>/export_report')
def export_class_report(class_id):
    """导出班级周报Excel"""
    conn = get_db_connection()
    students = conn.execute('''
        SELECT s.name, s.student_id, s.points, g.name as group_name,
               (SELECT content FROM student_evaluations WHERE student_id = s.id ORDER BY created_at DESC LIMIT 1) as last_eval
        FROM students s
        LEFT JOIN groups g ON s.group_id = g.id
        WHERE s.class_id = ?
        ORDER BY s.points DESC
    ''', (class_id,)).fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "班级学情报告"
    
    # 表头
    headers = ["排名", "姓名", "学号", "所属小组", "当前总积分", "老师评价"]
    ws.append(headers)
    
    for i, s in enumerate(students):
        ws.append([i+1, s['name'], s['student_id'], s['group_name'] or "-", s['points'], s['last_eval'] or "暂无评价"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    conn.close()
    
    filename = f"class_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename)

@app.route('/api/classes/<int:class_id>', methods=['DELETE'])
def delete_class(class_id):
    """彻底删除班级及其所有关联数据"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. 删除关联的评价
        cursor.execute('DELETE FROM student_evaluations WHERE student_id IN (SELECT id FROM students WHERE class_id = ?)', (class_id,))
        # 2. 删除关联的积分历史
        cursor.execute('DELETE FROM points_history WHERE student_id IN (SELECT id FROM students WHERE class_id = ?)', (class_id,))
        # 3. 删除关联的小组积分历史
        cursor.execute('DELETE FROM group_points_history WHERE group_id IN (SELECT id FROM groups WHERE class_id = ?)', (class_id,))
        # 4. 删除兑换记录
        cursor.execute('DELETE FROM redemptions WHERE student_id IN (SELECT id FROM students WHERE class_id = ?)', (class_id,))
        cursor.execute('DELETE FROM group_redemptions WHERE group_id IN (SELECT id FROM groups WHERE class_id = ?)', (class_id,))
        # 5. 删除学生
        cursor.execute('DELETE FROM students WHERE class_id = ?', (class_id,))
        # 6. 删除小组
        cursor.execute('DELETE FROM groups WHERE class_id = ?', (class_id,))
        # 7. 最后删除班级本身
        cursor.execute('DELETE FROM classes WHERE id = ?', (class_id,))
        
        conn.commit()
        conn.close()
        return jsonify({'message': '班级及关联数据已彻底清除'}), 200
    except Exception as e:
        print(f"删除班级失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/classes', methods=['GET'])
def get_classes():
    """获取所有班级"""
    try:
        if not check_and_init_tables():
            return jsonify({'error': '数据库初始化失败'}), 500
            
        conn = get_db_connection()
        classes = conn.execute('SELECT * FROM classes ORDER BY name').fetchall()
        conn.close()
        
        if not classes:
            print("数据库中没有班级")
            return jsonify([])
        
        result = []
        for c in classes:
            result.append({
                'id': c['id'],
                'name': c['name'],
                'teacher': c['teacher'] if 'teacher' in c.keys() else '',
                'created_at': c['created_at'] if 'created_at' in c.keys() else ''
            })
        
        print(f"返回 {len(result)} 个班级")
        return jsonify(result)
        
    except Exception as e:
        print(f"获取班级时发生错误: {e}")
        return jsonify([])

@app.route('/api/classes', methods=['POST'])
def create_class():
    """创建班级"""
    try:
        check_and_init_tables()
        
        data = request.json
        name = data.get('name')
        teacher = data.get('teacher', '')
        
        if not name:
            return jsonify({'error': '班级名称不能为空'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                'INSERT INTO classes (name, teacher) VALUES (?, ?)',
                (name, teacher)
            )
            class_id = cursor.lastrowid
            conn.commit()
            
            return jsonify({'id': class_id, 'name': name, 'teacher': teacher}), 201
        except sqlite3.IntegrityError:
            return jsonify({'error': '班级名称已存在'}), 400
        finally:
            conn.close()
    except Exception as e:
        print(f"创建班级失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/classes/<int:class_id>/students', methods=['GET'])
def get_students(class_id):
    """获取班级学生"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        students = conn.execute(
            '''SELECT s.*, g.name as group_name, g.color as group_color 
               FROM students s 
               LEFT JOIN groups g ON s.group_id = g.id 
               WHERE s.class_id = ? 
               ORDER BY s.name''',
            (class_id,)
        ).fetchall()
        conn.close()
        
        return jsonify([dict(s) for s in students])
    except Exception as e:
        print(f"获取学生失败: {e}")
        return jsonify([])

@app.route('/api/students', methods=['POST'])
def add_student():
    """添加学生"""
    try:
        check_and_init_tables()
        
        data = request.json
        class_id = data.get('class_id')
        name = data.get('name')
        student_id = data.get('student_id')
        group_id = data.get('group_id')
        
        if not all([class_id, name, student_id]):
            return jsonify({'error': '缺少必要参数'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                'INSERT INTO students (class_id, name, student_id, group_id) VALUES (?, ?, ?, ?)',
                (class_id, name, student_id, group_id)
            )
            new_student_id = cursor.lastrowid
            conn.commit()
            
            return jsonify({
                'id': new_student_id,
                'class_id': class_id,
                'name': name,
                'student_id': student_id,
                'group_id': group_id,
                'points': 0
            }), 201
        except sqlite3.IntegrityError:
            return jsonify({'error': '学号已存在'}), 400
        finally:
            conn.close()
    except Exception as e:
        print(f"添加学生失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/<int:student_id>', methods=['PUT'])
def update_student(student_id):
    """更新学生信息"""
    try:
        check_and_init_tables()
        
        data = request.json
        name = data.get('name')
        group_id = data.get('group_id')
        
        if not name:
            return jsonify({'error': '姓名不能为空'}), 400
        
        conn = get_db_connection()
        
        if group_id is not None:
            conn.execute(
                'UPDATE students SET name = ?, group_id = ? WHERE id = ?',
                (name, group_id, student_id)
            )
        else:
            conn.execute(
                'UPDATE students SET name = ? WHERE id = ?',
                (name, student_id)
            )
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': '学生信息更新成功'}), 200
    except Exception as e:
        print(f"更新学生失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    """删除学生"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
        conn.execute('DELETE FROM points_history WHERE student_id = ?', (student_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': '学生删除成功'}), 200
    except Exception as e:
        print(f"删除学生失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/<int:student_id>/points', methods=['POST'])
def update_points(student_id):
    """更新学生积分"""
    try:
        check_and_init_tables()
        
        data = request.json
        change_amount = data.get('change_amount', 0)
        reason = data.get('reason', '')
        teacher = data.get('teacher', '系统')
        
        if change_amount == 0:
            return jsonify({'error': '积分变化不能为0'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查学生是否存在
        student = conn.execute(
            'SELECT id FROM students WHERE id = ?', (student_id,)
        ).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'error': '学生不存在'}), 404
        
        # 更新学生积分
        cursor.execute(
            'UPDATE students SET points = points + ? WHERE id = ?',
            (change_amount, student_id)
        )
        
        # 记录积分历史
        cursor.execute(
            'INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, ?)',
            (student_id, change_amount, reason, teacher)
        )
        
        # 获取更新后的积分
        updated_student = conn.execute(
            'SELECT points FROM students WHERE id = ?',
            (student_id,)
        ).fetchone()
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': '积分更新成功',
            'new_points': updated_student['points']
        }), 200
    except Exception as e:
        print(f"更新积分失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/batch/points', methods=['POST'])
def batch_update_points():
    """批量更新学生积分（指定ID列表）"""
    try:
        check_and_init_tables()
        
        data = request.json
        student_ids = data.get('student_ids', [])
        change_amount = data.get('change_amount', 0)
        reason = data.get('reason', '')
        teacher = data.get('teacher', '系统')
        
        if not student_ids:
            return jsonify({'error': '未选择学生'}), 400
            
        if change_amount == 0:
            return jsonify({'error': '积分变化不能为0'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        success_count = 0
        
        # 批量处理
        for student_id in student_ids:
            # 检查学生是否存在
            student = conn.execute(
                'SELECT id FROM students WHERE id = ?', (student_id,)
            ).fetchone()
            
            if student:
                # 更新学生积分
                cursor.execute(
                    'UPDATE students SET points = points + ? WHERE id = ?',
                    (change_amount, student_id)
                )
                
                # 记录积分历史
                cursor.execute(
                    'INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, ?)',
                    (student_id, change_amount, reason, teacher)
                )
                success_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'成功更新 {success_count} 名学生的积分',
            'success_count': success_count
        }), 200
    except Exception as e:
        print(f"批量更新积分失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/<int:student_id>/history', methods=['GET'])
def get_points_history(student_id):
    """获取学生积分历史"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        history = conn.execute(
            '''SELECT ph.*, s.name as student_name 
               FROM points_history ph 
               JOIN students s ON ph.student_id = s.id 
               WHERE ph.student_id = ? 
               ORDER BY ph.created_at DESC''',
            (student_id,)
        ).fetchall()
        conn.close()
        
        return jsonify([dict(h) for h in history])
    except Exception as e:
        print(f"获取积分历史失败: {e}")
        return jsonify([])

@app.route('/api/classes/<int:class_id>/ranking', methods=['GET'])
def get_ranking(class_id):
    """获取积分排行榜（支持学生/分组，支持日期范围）"""
    try:
        check_and_init_tables()
        
        ranking_type = request.args.get('type', 'student') # student or group
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = get_db_connection()
        
        # 构建日期筛选条件
        date_filter = ""
        params = [class_id]
        if start_date and end_date:
            date_filter = "AND date(ph.created_at) BETWEEN ? AND ?"
            params.extend([start_date, end_date])
        
        if ranking_type == 'student':
            if date_filter:
                # 按历史记录统计增量
                sql = f'''
                    SELECT s.id, s.name, s.student_id, 
                           COALESCE(SUM(ph.change_amount), 0) as points,
                           g.name as group_name, g.color as group_color
                    FROM students s
                    LEFT JOIN points_history ph ON s.id = ph.student_id
                    LEFT JOIN groups g ON s.group_id = g.id
                    WHERE s.class_id = ? {date_filter}
                    GROUP BY s.id
                    ORDER BY points DESC, s.name ASC
                '''
            else:
                # 直接查询当前总分
                sql = '''
                    SELECT s.id, s.name, s.student_id, s.points,
                           g.name as group_name, g.color as group_color
                    FROM students s
                    LEFT JOIN groups g ON s.group_id = g.id
                    WHERE s.class_id = ?
                    ORDER BY s.points DESC, s.name ASC
                '''
                params = [class_id] # 重置params
                
        else: # type == 'group'
            if date_filter:
                # 统计组内学生历史记录总和
                sql = f'''
                    SELECT g.id, g.name, g.color,
                           COALESCE(SUM(ph.change_amount), 0) as points,
                           COUNT(DISTINCT s.id) as student_count
                    FROM groups g
                    LEFT JOIN students s ON g.id = s.group_id
                    LEFT JOIN points_history ph ON s.id = ph.student_id
                    WHERE g.class_id = ? {date_filter}
                    GROUP BY g.id
                    ORDER BY points DESC, g.name ASC
                '''
            else:
                # 统计当前总分
                sql = '''
                    SELECT g.id, g.name, g.color,
                           COALESCE(SUM(s.points), 0) as points,
                           COUNT(s.id) as student_count
                    FROM groups g
                    LEFT JOIN students s ON g.id = s.group_id
                    WHERE g.class_id = ?
                    GROUP BY g.id
                    ORDER BY points DESC, g.name ASC
                '''
                params = [class_id]

        ranking = conn.execute(sql, params).fetchall()
        conn.close()
        
        # 处理排名
        result = []
        for i, r in enumerate(ranking):
            item = dict(r)
            item['rank'] = i + 1
            result.append(item)
            
        return jsonify(result)
    except Exception as e:
        print(f"获取排行榜失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/classes/<int:class_id>/ranking/chart', methods=['GET'])
def get_ranking_chart(class_id):
    """获取图表数据"""
    try:
        check_and_init_tables()
        # 这里暂不使用日期筛选，默认展示最近趋势
        
        conn = get_db_connection()
        
        # 1. 积分分布 (前5名学生)
        top_students = conn.execute('''
            SELECT name, points FROM students 
            WHERE class_id = ? 
            ORDER BY points DESC LIMIT 5
        ''', (class_id,)).fetchall()
        
        distribution = {
            'labels': [s['name'] for s in top_students],
            'data': [s['points'] for s in top_students]
        }
        
        # 2. 积分趋势 (最近7天全班总积分变化)
        trend_sql = '''
            SELECT date(ph.created_at) as date, SUM(change_amount) as total_change
            FROM points_history ph
            JOIN students s ON ph.student_id = s.id
            WHERE s.class_id = ?
            GROUP BY date(ph.created_at)
            ORDER BY date(ph.created_at) DESC
            LIMIT 7
        '''
        trend_data = conn.execute(trend_sql, (class_id,)).fetchall()
        
        # 整理趋势数据（按日期升序）
        trend_data.reverse()
        trend = {
            'labels': [t['date'] for t in trend_data],
            'data': [t['total_change'] for t in trend_data]
        }
        
        # 3. 分组对比 (Bar Chart) - 各组平均分
        groups = conn.execute('''
            SELECT g.name, AVG(s.points) as avg_points
            FROM groups g
            LEFT JOIN students s ON g.id = s.group_id
            WHERE g.class_id = ?
            GROUP BY g.id
            ORDER BY avg_points DESC
        ''', (class_id,)).fetchall()
        
        group_comp = {
            'labels': [g['name'] for g in groups],
            'data': [round(g['avg_points'] or 0, 1) for g in groups]
        }
        
        conn.close()
        
        return jsonify({
            'distribution': distribution,
            'trend': trend,
            'group_comp': group_comp
        })
        
    except Exception as e:
        print(f"获取图表数据失败: {e}")
        return jsonify({'error': str(e)}), 500

# ============ 分组管理 API ============

@app.route('/api/groups', methods=['GET'])
def get_groups():
    """获取所有分组（可筛选班级）"""
    try:
        check_and_init_tables()
        
        class_id = request.args.get('class_id')
        
        conn = get_db_connection()
        if class_id:
            groups = conn.execute(
                '''SELECT g.*, 
                          COUNT(s.id) as student_count,
                          SUM(s.points) as total_points,
                          AVG(s.points) as avg_points
                   FROM groups g
                   LEFT JOIN students s ON g.id = s.group_id
                   WHERE g.class_id = ?
                   GROUP BY g.id
                   ORDER BY g.name''',
                (class_id,)
            ).fetchall()
        else:
            groups = conn.execute(
                '''SELECT g.*, 
                          COUNT(s.id) as student_count,
                          SUM(s.points) as total_points,
                          AVG(s.points) as avg_points
                   FROM groups g
                   LEFT JOIN students s ON g.id = s.group_id
                   GROUP BY g.id
                   ORDER BY g.name'''
            ).fetchall()
        
        conn.close()
        
        result = []
        for g in groups:
            group_dict = dict(g)
            group_dict['avg_points'] = round(group_dict['avg_points'] or 0, 2)
            result.append(group_dict)
        
        return jsonify(result)
    except Exception as e:
        print(f"获取分组失败: {e}")
        return jsonify([])

@app.route('/api/groups', methods=['POST'])
def create_group():
    """创建分组"""
    try:
        check_and_init_tables()
        
        data = request.json
        class_id = data.get('class_id')
        name = data.get('name')
        color = data.get('color', '#667eea')
        
        if not class_id or not name:
            return jsonify({'error': '缺少必要参数'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                'INSERT INTO groups (class_id, name, color) VALUES (?, ?, ?)',
                (class_id, name, color)
            )
            group_id = cursor.lastrowid
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'id': group_id,
                'class_id': class_id,
                'name': name,
                'color': color,
                'student_count': 0,
                'total_points': 0,
                'avg_points': 0
            }), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': '分组名称已存在'}), 400
    except Exception as e:
        print(f"创建分组失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups/<int:group_id>', methods=['PUT'])
def update_group(group_id):
    """更新分组"""
    try:
        check_and_init_tables()
        
        data = request.json
        name = data.get('name')
        color = data.get('color')
        
        if not name and not color:
            return jsonify({'error': '没有要更新的数据'}), 400
        
        conn = get_db_connection()
        if name:
            conn.execute('UPDATE groups SET name = ? WHERE id = ?', (name, group_id))
        if color:
            conn.execute('UPDATE groups SET color = ? WHERE id = ?', (color, group_id))
        
        conn.commit()
        conn.close()
        return jsonify({'message': '分组更新成功'}), 200
    except Exception as e:
        print(f"更新分组失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups/<int:group_id>', methods=['DELETE'])
def delete_group(group_id):
    """删除分组"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        
        # 将属于该分组的学生设为无分组
        conn.execute('UPDATE students SET group_id = NULL WHERE group_id = ?', (group_id,))
        
        # 删除分组
        conn.execute('DELETE FROM groups WHERE id = ?', (group_id,))
        
        # 删除分组积分历史
        conn.execute('DELETE FROM group_points_history WHERE group_id = ?', (group_id,))
        
        conn.commit()
        conn.close()
        return jsonify({'message': '分组删除成功'}), 200
    except Exception as e:
        print(f"删除分组失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups/<int:group_id>/students', methods=['GET'])
def get_group_students(group_id):
    """获取分组内的学生"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        students = conn.execute(
            'SELECT * FROM students WHERE group_id = ? ORDER BY name',
            (group_id,)
        ).fetchall()
        conn.close()
        return jsonify([dict(s) for s in students])
    except Exception as e:
        print(f"获取分组学生失败: {e}")
        return jsonify([])

@app.route('/api/groups/<int:group_id>/stats', methods=['GET'])
def get_group_stats(group_id):
    """获取分组统计信息"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        
        # 分组信息
        group = conn.execute(
            'SELECT * FROM groups WHERE id = ?', (group_id,)
        ).fetchone()
        
        if not group:
            conn.close()
            return jsonify({'error': '分组不存在'}), 404
        
        # 分组学生数
        student_count = conn.execute(
            'SELECT COUNT(*) as count FROM students WHERE group_id = ?',
            (group_id,)
        ).fetchone()['count']
        
        # 分组平均积分
        avg_points_result = conn.execute(
            'SELECT AVG(points) as avg FROM students WHERE group_id = ?',
            (group_id,)
        ).fetchone()
        avg_points = avg_points_result['avg'] if avg_points_result['avg'] is not None else 0
        
        # 分组总积分
        total_points_result = conn.execute(
            'SELECT SUM(points) as total FROM students WHERE group_id = ?',
            (group_id,)
        ).fetchone()
        total_points = total_points_result['total'] if total_points_result['total'] is not None else 0
        
        conn.close()
        
        return jsonify({
            'group': dict(group),
            'student_count': student_count,
            'avg_points': round(avg_points, 2),
            'total_points': total_points
        })
    except Exception as e:
        print(f"获取分组统计失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups/<int:group_id>/points', methods=['POST'])
def update_group_points(group_id):
    """批量更新分组学生积分"""
    try:
        check_and_init_tables()
        
        data = request.json
        change_amount = data.get('change_amount', 0)
        reason = data.get('reason', '')
        teacher = data.get('teacher', '系统')
        
        if change_amount == 0:
            return jsonify({'error': '积分变化不能为0'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取分组内的所有学生
        students = conn.execute(
            'SELECT id FROM students WHERE group_id = ?',
            (group_id,)
        ).fetchall()
        
        if not students:
            conn.close()
            return jsonify({'error': '分组内没有学生'}), 400
        
        updated_count = 0
        for student in students:
            student_id = student['id']
            
            # 更新学生积分
            cursor.execute(
                'UPDATE students SET points = points + ? WHERE id = ?',
                (change_amount, student_id)
            )
            
            # 记录个人积分历史
            cursor.execute(
                '''INSERT INTO points_history 
                   (student_id, change_amount, reason, teacher) 
                   VALUES (?, ?, ?, ?)''',
                (student_id, change_amount, reason, teacher)
            )
            
            updated_count += 1
        
        # 记录分组积分历史
        cursor.execute(
            '''INSERT INTO group_points_history 
               (group_id, change_amount, reason, teacher) 
               VALUES (?, ?, ?, ?)''',
            (group_id, change_amount * updated_count, reason, teacher)
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'成功为{updated_count}名学生更新积分',
            'updated_count': updated_count,
            'total_change': change_amount * updated_count
        }), 200
    except Exception as e:
        print(f"批量更新分组积分失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/classes/<int:class_id>/group-ranking', methods=['GET'])
def get_group_ranking(class_id):
    """获取班级内分组积分排行"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        
        # 检查班级是否存在
        class_exists = conn.execute(
            'SELECT id FROM classes WHERE id = ?', (class_id,)
        ).fetchone()
        
        if not class_exists:
            conn.close()
            return jsonify([])
        
        groups = conn.execute(
            '''SELECT g.*, 
                      COUNT(s.id) as student_count,
                      SUM(s.points) as total_points,
                      AVG(s.points) as avg_points
               FROM groups g
               LEFT JOIN students s ON g.id = s.group_id
               WHERE g.class_id = ?
               GROUP BY g.id
               ORDER BY total_points DESC, avg_points DESC''',
            (class_id,)
        ).fetchall()
        
        conn.close()
        
        result = []
        for i, group in enumerate(groups):
            avg_points = group['avg_points'] if group['avg_points'] is not None else 0
            result.append({
                'id': group['id'],
                'class_id': group['class_id'],
                'name': group['name'],
                'color': group['color'],
                'student_count': group['student_count'] or 0,
                'total_points': group['total_points'] or 0,
                'avg_points': round(avg_points, 2),
                'rank': i + 1
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"获取分组排行失败: {e}")
        return jsonify([])

@app.route('/api/students/<int:student_id>/group', methods=['PUT'])
def update_student_group(student_id):
    """更新学生所在分组"""
    try:
        check_and_init_tables()
        
        data = request.json
        group_id = data.get('group_id')  # null 表示无分组
        
        conn = get_db_connection()
        
        # 检查学生是否存在
        student = conn.execute(
            'SELECT id FROM students WHERE id = ?', (student_id,)
        ).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'error': '学生不存在'}), 404
        
        # 如果提供了group_id，检查分组是否存在
        if group_id is not None:
            group = conn.execute(
                'SELECT id FROM groups WHERE id = ?', (group_id,)
            ).fetchone()
            
            if not group:
                conn.close()
                return jsonify({'error': '分组不存在'}), 404
        
        conn.execute(
            'UPDATE students SET group_id = ? WHERE id = ?',
            (group_id, student_id)
        )
        conn.commit()
        
        # 获取更新后的学生信息
        updated_student = conn.execute(
            '''SELECT s.*, g.name as group_name, g.color as group_color 
               FROM students s 
               LEFT JOIN groups g ON s.group_id = g.id 
               WHERE s.id = ?''',
            (student_id,)
        ).fetchone()
        
        conn.close()
        return jsonify(dict(updated_student)), 200
    except Exception as e:
        print(f"更新学生分组失败: {e}")
        return jsonify({'error': str(e)}), 500

# ============ Excel导入导出API ============

@app.route('/api/students/import', methods=['POST'])
def import_students():
    """从Excel导入学生"""
    try:
        check_and_init_tables()
        
        # 检查文件是否存在
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        # 检查文件格式
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '只支持Excel文件 (.xlsx, .xls)'}), 400
        
        # 获取班级ID
        class_id = request.form.get('class_id')
        if not class_id:
            return jsonify({'error': '请选择班级'}), 400
        
        # 自动创建分组选项
        auto_create_group = request.form.get('auto_create_group', 'false') == 'true'
        
        # 将文件内容读取到内存
        file_content = file.read()
        
        try:
            # 使用 openpyxl 读取
            wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
            sheet = wb.active
            
            # 获取表头
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
                
            # 检查必要的列
            required_columns = ['姓名', '学号']
            col_indices = {}
            for col in required_columns:
                if col not in headers:
                    return jsonify({'error': f'Excel文件必须包含"{col}"列'}), 400
                col_indices[col] = headers.index(col)
            
            # 获取可选列的索引
            if '分组' in headers:
                col_indices['分组'] = headers.index('分组')
            if '初始积分' in headers:
                col_indices['初始积分'] = headers.index('初始积分')

            # 处理数据
            success_count = 0
            fail_count = 0
            errors = []
            
            conn = get_db_connection()
            
            # 从第二行开始遍历
            row_idx = 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    # 获取数据，处理可能为空的情况
                    def get_val(idx):
                        if idx < len(row) and row[idx] is not None:
                            return row[idx]
                        return ''

                    name = str(get_val(col_indices['姓名'])).strip()
                    student_id = str(get_val(col_indices['学号'])).strip()
                    
                    # 检查必填字段
                    if not name or not student_id:
                        # 只有当整行都为空时才忽略，否则报错
                        if not any(str(cell).strip() for cell in row if cell is not None):
                            continue
                        errors.append(f"第{row_idx}行: 姓名和学号不能为空")
                        fail_count += 1
                        row_idx += 1
                        continue
                    
                    # 检查学号是否已存在
                    existing = conn.execute(
                        'SELECT id FROM students WHERE student_id = ?',
                        (student_id,)
                    ).fetchone()
                    
                    if existing:
                        errors.append(f"第{row_idx}行: 学号 {student_id} 已存在")
                        fail_count += 1
                        row_idx += 1
                        continue
                    
                    # 可选字段 - 分组
                    group_id = None
                    if '分组' in col_indices:
                        group_name = str(get_val(col_indices['分组'])).strip()
                        if group_name and group_name != 'None':
                            # 根据分组名称查找分组ID
                            group = conn.execute(
                                'SELECT id FROM groups WHERE name = ? AND class_id = ?',
                                (group_name, class_id)
                            ).fetchone()
                            if group:
                                group_id = group['id']
                            elif auto_create_group:
                                # 自动创建分组
                                cursor = conn.cursor()
                                try:
                                    cursor.execute(
                                        'INSERT INTO groups (class_id, name) VALUES (?, ?)',
                                        (class_id, group_name)
                                    )
                                    group_id = cursor.lastrowid
                                    conn.commit()
                                except sqlite3.IntegrityError:
                                    # 分组已存在（并发情况）
                                    group = conn.execute(
                                        'SELECT id FROM groups WHERE name = ? AND class_id = ?',
                                        (group_name, class_id)
                                    ).fetchone()
                                    if group:
                                        group_id = group['id']
                                    else:
                                        errors.append(f"第{row_idx}行: 无法创建分组 '{group_name}'")
                                        fail_count += 1
                                        row_idx += 1
                                        continue
                            else:
                                errors.append(f"第{row_idx}行: 分组 '{group_name}' 不存在")
                                fail_count += 1
                                row_idx += 1
                                continue
                    
                    # 初始积分（可选）
                    initial_points = 0
                    if '初始积分' in col_indices:
                        try:
                            val = get_val(col_indices['初始积分'])
                            if val:
                                initial_points = int(float(val)) # 处理可能是float的情况
                        except:
                            initial_points = 0
                    
                    # 插入学生数据
                    cursor = conn.cursor()
                    cursor.execute(
                        'INSERT INTO students (class_id, group_id, name, student_id, points) VALUES (?, ?, ?, ?, ?)',
                        (class_id, group_id, name, student_id, initial_points)
                    )
                    new_student_id = cursor.lastrowid
                    
                    # 如果有初始积分，记录历史
                    if initial_points != 0:
                        cursor.execute(
                            'INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, ?)',
                            (new_student_id, initial_points, '初始积分', '系统导入')
                        )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"第{row_idx}行: {str(e)}")
                    fail_count += 1
                
                row_idx += 1
                
        except Exception as e:
             return jsonify({'error': f'无法读取Excel文件: {str(e)}'}), 400
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'导入完成，成功 {success_count} 条，失败 {fail_count} 条',
            'success_count': success_count,
            'fail_count': fail_count,
            'errors': errors[:10]  # 只返回前10个错误
        }), 200
        
    except Exception as e:
        print(f"导入学生失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/template', methods=['GET'])
def download_template():
    """下载Excel模板"""
    try:
        wb = Workbook()
        
        # --- 学生名单 Sheet ---
        ws1 = wb.active
        ws1.title = '学生名单'
        
        # 写入表头
        headers = ['姓名', '学号', '分组', '初始积分']
        ws1.append(headers)
        
        # 写入示例数据
        data = [
            ['张三', '2023001', '第一组', 100],
            ['李四', '2023002', '第二组', 80],
            ['王五', '2023003', '第一组', 90]
        ]
        for row in data:
            ws1.append(row)

        # 设置列宽
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 15
            
        # --- 使用说明 Sheet ---
        ws2 = wb.create_sheet(title='使用说明')
        
        # 写入表头
        ws2.append(['列名', '说明', '示例'])
        
        # 写入说明
        instructions = [
            ['姓名', '学生姓名（必填）', '张三'],
            ['学号', '学生学号，必须唯一（必填）', '2023001'],
            ['分组', '所在分组名称（选填，如果分组不存在会自动创建）', '第一组'],
            ['初始积分', '初始积分，默认为0（选填）', '100']
        ]
        for row in instructions:
            ws2.append(row)
            
        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 使用纯英文文件名避免编码问题
        filename = 'student_import_template.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"下载模板失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/export', methods=['GET'])
def export_students():
    """导出学生列表到Excel"""
    try:
        check_and_init_tables()
        
        class_id = request.args.get('class_id')
        if not class_id:
            return jsonify({'error': '请选择班级'}), 400
        
        conn = get_db_connection()
        
        # 获取班级信息
        class_info = conn.execute(
            'SELECT name FROM classes WHERE id = ?', (class_id,)
        ).fetchone()
        
        if not class_info:
            conn.close()
            return jsonify({'error': '班级不存在'}), 404
        
        # 获取学生数据
        students = conn.execute(
            '''SELECT s.name, s.student_id, s.points, g.name as group_name, 
                      s.created_at, s.id
               FROM students s 
               LEFT JOIN groups g ON s.group_id = g.id 
               WHERE s.class_id = ?
               ORDER BY s.name''',
            (class_id,)
        ).fetchall()
        
        # 获取班级分组
        groups = conn.execute(
            'SELECT id, name FROM groups WHERE class_id = ? ORDER BY name',
            (class_id,)
        ).fetchall()
        
        conn.close()
        
        # 创建Excel文件
        wb = Workbook()
        
        # --- Students Sheet ---
        ws1 = wb.active
        ws1.title = 'Students'
        
        # 写入表头
        ws1.append(['Name', 'StudentID', 'Group', 'Points', 'JoinDate'])
        
        # 写入数据
        for student in students:
            ws1.append([
                student['name'],
                student['student_id'],
                student['group_name'] or '',
                student['points'],
                student['created_at']
            ])
            
        # 设置列宽
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 10
        ws1.column_dimensions['E'].width = 20

        # --- Groups Sheet ---
        if groups:
            ws2 = wb.create_sheet(title='Groups')
            ws2.append(['GroupID', 'GroupName'])
            for group in groups:
                ws2.append([group['id'], group['name']])
                
            ws2.column_dimensions['B'].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 使用纯英文文件名避免编码问题
        filename = f"students_export_{class_id}_{int(time.time())}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出学生失败: {e}")
        return jsonify({'error': str(e)}), 500

# ============ 班级统计 API ============

@app.route('/api/classes/<int:class_id>/stats', methods=['GET'])
def get_class_stats(class_id):
    """获取班级统计信息"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        
        # 检查班级是否存在
        class_info = conn.execute(
            'SELECT * FROM classes WHERE id = ?', (class_id,)
        ).fetchone()
        
        if not class_info:
            conn.close()
            return jsonify({'error': '班级不存在'}), 404
        
        # 学生总数
        total_students_result = conn.execute(
            'SELECT COUNT(*) as count FROM students WHERE class_id = ?',
            (class_id,)
        ).fetchone()
        total_students = total_students_result['count']
        
        # 平均积分
        avg_points_result = conn.execute(
            'SELECT AVG(points) as avg FROM students WHERE class_id = ?',
            (class_id,)
        ).fetchone()
        avg_points = avg_points_result['avg'] if avg_points_result['avg'] is not None else 0
        
        # 最高积分
        max_points_result = conn.execute(
            'SELECT MAX(points) as max FROM students WHERE class_id = ?',
            (class_id,)
        ).fetchone()
        max_points = max_points_result['max'] if max_points_result['max'] is not None else 0
        
        # 最低积分
        min_points_result = conn.execute(
            'SELECT MIN(points) as min FROM students WHERE class_id = ?',
            (class_id,)
        ).fetchone()
        min_points = min_points_result['min'] if min_points_result['min'] is not None else 0
        
        # 分组统计
        group_stats_result = conn.execute(
            '''SELECT COUNT(DISTINCT g.id) as group_count,
                      AVG(group_avg.avg_points) as groups_avg_points
               FROM groups g
               LEFT JOIN (
                   SELECT g.id, AVG(s.points) as avg_points
                   FROM groups g
                   LEFT JOIN students s ON g.id = s.group_id
                   WHERE g.class_id = ?
                   GROUP BY g.id
               ) group_avg ON g.id = group_avg.id
               WHERE g.class_id = ?''',
            (class_id, class_id)
        ).fetchone()
        
        group_count = group_stats_result['group_count'] or 0
        groups_avg_points = group_stats_result['groups_avg_points'] or 0
        
        # 最近积分变动
        recent_changes = conn.execute(
            '''SELECT ph.*, s.name as student_name 
               FROM points_history ph 
               JOIN students s ON ph.student_id = s.id 
               WHERE s.class_id = ? 
               ORDER BY ph.created_at DESC LIMIT 10''',
            (class_id,)
        ).fetchall()
        
        conn.close()
        
        return jsonify({
            'class_info': dict(class_info),
            'total_students': total_students,
            'avg_points': round(avg_points, 2),
            'max_points': max_points,
            'min_points': min_points,
            'group_count': group_count,
            'groups_avg_points': round(groups_avg_points, 2),
            'recent_changes': [dict(c) for c in recent_changes]
        })
    except Exception as e:
        print(f"获取班级统计失败: {e}")
        return jsonify({'error': str(e)}), 500

# ============ 模板路由 ============

@app.route('/students')
def students_page():
    """学生管理页面"""
    return render_template('students.html')

@app.route('/points')
def points_page():
    """积分管理页面"""
    return render_template('points.html')

@app.route('/ranking')
def ranking_page():
    """排行榜页面"""
    return render_template('ranking.html')

@app.route('/groups')
def groups_page():
    """分组管理页面"""
    return render_template('groups.html')

@app.route('/api/test')
def test_api():
    """测试API端点"""
    try:
        check_and_init_tables()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查所有表
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        table_list = [t[0] for t in tables]
        
        # 统计数据
        stats = {}
        for table in table_list:
            cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
            count = cursor.fetchone()['count']
            stats[table] = count
        
        conn.close()
        
        return jsonify({
            'status': 'ok',
            'message': 'API服务器正常运行',
            'timestamp': datetime.now().isoformat(),
            'database_path': app.config['DATABASE_PATH'],
            'tables': table_list,
            'stats': stats
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/test')
def test_page():
    """测试页面"""
    return """
    <!DOCTYPE html>
    <html>
    <head><title>测试页</title></head>
    <body>
        <h1>班级积分管理系统测试页面</h1>
        <p><a href="/api/classes" target="_blank">测试班级API</a></p>
        <p><a href="/api/groups" target="_blank">测试分组API</a></p>
        <p><a href="/api/students/template" target="_blank">下载Excel模板</a></p>
        <p><a href="/api/test" target="_blank">测试系统状态</a></p>
        <p><a href="/" target="_blank">返回主页面</a></p>
    </body>
    </html>
    """

# ============ 奖品与兑换 API ============

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}

@app.route('/api/rewards', methods=['GET'])
def get_rewards():
    """获取所有奖品"""
    try:
        check_and_init_tables()
        conn = get_db_connection()
        rewards = conn.execute('SELECT * FROM rewards ORDER BY created_at DESC').fetchall()
        conn.close()
        return jsonify([dict(r) for r in rewards])
    except Exception as e:
        print(f"获取奖品失败: {e}")
        return jsonify([])

@app.route('/api/rewards', methods=['POST'])
def add_reward():
    """添加奖品"""
    try:
        check_and_init_tables()
        
        name = request.form.get('name')
        description = request.form.get('description')
        points_cost = request.form.get('points_cost')
        stock = request.form.get('stock', 10)
        
        if not name or not points_cost:
            return jsonify({'error': '名称和积分是必填项'}), 400
            
        image_path = ''
        if 'image' in request.files:
            file = request.files['image']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # 添加时间戳防止重名
                filename = f"{int(time.time())}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = f"/static/uploads/{filename}"
        
        conn = get_db_connection()
        conn.execute(
            'INSERT INTO rewards (name, description, points_cost, image_path, stock) VALUES (?, ?, ?, ?, ?)',
            (name, description, int(points_cost), image_path, int(stock))
        )
        conn.commit()
        conn.close()
        
        return jsonify({'message': '奖品添加成功'}), 201
    except Exception as e:
        print(f"添加奖品失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/rewards/<int:reward_id>', methods=['DELETE'])
def delete_reward(reward_id):
    """删除奖品"""
    try:
        check_and_init_tables()
        conn = get_db_connection()
        conn.execute('DELETE FROM rewards WHERE id = ?', (reward_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': '奖品删除成功'}), 200
    except Exception as e:
        print(f"删除奖品失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students/<int:student_id>/rewards', methods=['GET'])
def get_student_rewards(student_id):
    """获取学生的兑换状态"""
    try:
        check_and_init_tables()
        conn = get_db_connection()
        
        # 获取所有奖品
        rewards = conn.execute('SELECT * FROM rewards ORDER BY points_cost ASC').fetchall()
        
        # 获取该学生的兑换记录
        redemptions = conn.execute(
            'SELECT reward_id, redeemed_at FROM redemptions WHERE student_id = ?', 
            (student_id,)
        ).fetchall()
        
        redeemed_ids = {r['reward_id'] for r in redemptions}
        
        result = []
        for r in rewards:
            item = dict(r)
            item['redeemed'] = r['id'] in redeemed_ids
            result.append(item)
            
        conn.close()
        return jsonify(result)
    except Exception as e:
        print(f"获取学生奖品状态失败: {e}")
        return jsonify([])

@app.route('/api/students/<int:student_id>/redeem', methods=['POST'])
def redeem_reward(student_id):
    """兑换奖品"""
    try:
        check_and_init_tables()
        data = request.json
        reward_id = data.get('reward_id')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. 检查学生是否存在及积分是否足够
        student = conn.execute('SELECT points FROM students WHERE id = ?', (student_id,)).fetchone()
        if not student:
            return jsonify({'error': '学生不存在'}), 404
            
        # 2. 检查奖品是否存在
        reward = conn.execute('SELECT points_cost, name, stock FROM rewards WHERE id = ?', (reward_id,)).fetchone()
        if not reward:
            return jsonify({'error': '奖品不存在'}), 404
        
        if reward['stock'] <= 0:
            return jsonify({'error': '奖品库存不足'}), 400
            
        # 3. 检查是否已兑换
        existing = conn.execute(
            'SELECT id FROM redemptions WHERE student_id = ? AND reward_id = ?',
            (student_id, reward_id)
        ).fetchone()
        
        if existing:
            return jsonify({'error': '该奖品已兑换'}), 400
            
        # 4. 检查积分是否足够
        if student['points'] < reward['points_cost']:
            return jsonify({'error': f'积分不足，需要 {reward["points_cost"]} 分'}), 400
            
        # 5. 执行兑换：扣分 + 扣库存 + 记录
        cursor.execute(
            'UPDATE students SET points = points - ? WHERE id = ?',
            (reward["points_cost"], student_id)
        )
        
        cursor.execute(
            'UPDATE rewards SET stock = stock - 1 WHERE id = ?',
            (reward_id,)
        )
        
        cursor.execute(
            'INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, ?)',
            (student_id, -reward["points_cost"], f'兑换奖品：{reward["name"]}', '系统')
        )
        
        cursor.execute(
            'INSERT INTO redemptions (student_id, reward_id) VALUES (?, ?)',
            (student_id, reward_id)
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': '兑换成功'}), 200
        
    except Exception as e:
        print(f"兑换失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups/<int:group_id>/redeem', methods=['POST'])
def redeem_group_reward(group_id):
    """小组兑换奖品"""
    try:
        check_and_init_tables()
        data = request.json
        reward_id = data.get('reward_id')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. 检查奖品是否存在
        reward = conn.execute('SELECT points_cost, name, stock FROM rewards WHERE id = ?', (reward_id,)).fetchone()
        if not reward:
            conn.close()
            return jsonify({'error': '奖品不存在'}), 404
            
        if reward['stock'] <= 0:
            conn.close()
            return jsonify({'error': '奖品库存不足'}), 400
            
        # 2. 获取小组学生
        students = conn.execute('SELECT id, name, points FROM students WHERE group_id = ?', (group_id,)).fetchall()
        if not students:
            conn.close()
            return jsonify({'error': '该小组没有学生'}), 400
            
        student_count = len(students)
        cost_per_student = reward['points_cost'] // student_count
        
        if cost_per_student == 0:
             # 如果奖品积分小于人数，至少每人扣1分？或者不允许？
             # 这里设为每人扣1分
             cost_per_student = 1

        total_cost = cost_per_student * student_count
        
        # 3. 检查每个学生积分是否足够
        for s in students:
            if s['points'] < cost_per_student:
                conn.close()
                return jsonify({'error': f'学生 {s["name"]} 积分不足 (需要 {cost_per_student} 分)'}), 400
        
        # 4. 执行扣分
        for s in students:
            cursor.execute(
                'UPDATE students SET points = points - ? WHERE id = ?',
                (cost_per_student, s['id'])
            )
            cursor.execute(
                'INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, ?)',
                (s['id'], -cost_per_student, f'小组兑换：{reward["name"]}', '系统')
            )
        
        # 5. 扣减库存
        cursor.execute(
            'UPDATE rewards SET stock = stock - 1 WHERE id = ?',
            (reward_id,)
        )
            
        # 6. 记录小组兑换
        cursor.execute(
            'INSERT INTO group_redemptions (group_id, reward_id) VALUES (?, ?)',
            (group_id, reward_id)
        )
        
        # 7. 记录小组积分变动历史 (用于统计)
        cursor.execute(
            'INSERT INTO group_points_history (group_id, change_amount, reason, teacher) VALUES (?, ?, ?, ?)',
            (group_id, -total_cost, f'兑换奖品：{reward["name"]}', '系统')
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': f'兑换成功，每位成员扣除 {cost_per_student} 积分'}), 200
        
    except Exception as e:
        print(f"小组兑换失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/redemptions/history', methods=['GET'])
def get_redemption_history():
    """获取兑换记录"""
    try:
        check_and_init_tables()
        student_id = request.args.get('student_id')
        group_id = request.args.get('group_id')
        
        conn = get_db_connection()
        
        if student_id:
            sql = '''
                SELECT r.name as reward_name, r.image_path, rd.redeemed_at, r.points_cost
                FROM redemptions rd
                JOIN rewards r ON rd.reward_id = r.id
                WHERE rd.student_id = ?
                ORDER BY rd.redeemed_at DESC
            '''
            records = conn.execute(sql, (student_id,)).fetchall()
        elif group_id:
            sql = '''
                SELECT r.name as reward_name, r.image_path, gr.redeemed_at, r.points_cost
                FROM group_redemptions gr
                JOIN rewards r ON gr.reward_id = r.id
                WHERE gr.group_id = ?
                ORDER BY gr.redeemed_at DESC
            '''
            records = conn.execute(sql, (group_id,)).fetchall()
        else:
            records = []
            
        conn.close()
        return jsonify([dict(r) for r in records])
    except Exception as e:
        print(f"获取兑换记录失败: {e}")
        return jsonify([])

if __name__ == '__main__':
    print("=" * 60)
    print("班级积分管理系统")
    print("=" * 60)
    
    # 检查数据库路径
    print(f"数据库路径: {app.config['DATABASE_PATH']}")
    
    # 确保data目录存在
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    if not os.path.exists(data_dir):
        print(f"创建数据目录: {data_dir}")
        os.makedirs(data_dir, exist_ok=True)
    
    # 初始化数据库
    print("正在初始化数据库...")
    success, message = init_db()
    if success:
        print(message)
    else:
        print(f"数据库初始化失败: {message}")
    
    # 创建示例数据
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查是否有班级
        cursor.execute("SELECT COUNT(*) as count FROM classes")
        class_count = cursor.fetchone()['count']
        print(f"班级数量: {class_count}")
        
        if class_count == 0:
            print("数据库已初始化。")
            print("当前系统中没有班级数据，请在网页端创建新班级。")
        
        # 显示当前数据统计
        print("\n当前数据统计:")
        cursor.execute("SELECT COUNT(*) as count FROM groups")
        group_count = cursor.fetchone()['count']
        print(f"分组数量: {group_count}")
        
        cursor.execute("SELECT COUNT(*) as count FROM students")
        student_count = cursor.fetchone()['count']
        print(f"学生数量: {student_count}")
        
        conn.close()
        
    except Exception as e:
        print(f"检查数据库状态时出错: {e}")
    
    print("=" * 60)
    
    # 获取本机IP
    try:
        # 获取本机主机名
        hostname = socket.gethostname()
        # 获取本机IP
        local_ip = socket.gethostbyname(hostname)
        
        # 尝试通过连接外部地址获取真实的局域网IP (更准确)
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            # 不需要真的连接
            s.connect(('8.8.8.8', 80))
            local_ip = s.getsockname()[0]
        except Exception:
            pass
        finally:
            s.close()
    except:
        local_ip = '127.0.0.1'

    port = 5001
    url = f"http://localhost:{port}"
    mobile_url = f"http://{local_ip}:{port}"

    print(f"系统启动成功！")
    print("-" * 30)
    print(f"【电脑端访问】: {url}")
    print(f"【手机端访问】: {mobile_url}")
    print("-" * 30)
    print("提示：请确保手机和电脑连接在同一个 Wi-Fi 下。")
    print("=" * 60)
    
    # 自动打开浏览器
    if not os.environ.get("WERKZEUG_RUN_MAIN"): # 防止重载时重复打开
        try:
            webbrowser.open_new(url)
        except:
            pass
    
    # 启动服务器
    app.run(host='0.0.0.0', port=port, debug=False) # 生产环境建议关闭 debug