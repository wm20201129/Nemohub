from flask import Flask, render_template, jsonify, request, send_file, make_response, session, redirect, url_for, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import sqlite3, json, os, io, sys, re, time, threading, datetime, socket, webbrowser
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pyngrok import ngrok, conf

# --- 1. 配置与路径 ---
class Config:
    SECRET_KEY = 'class-points-manager-secret-2025'
    if getattr(sys, 'frozen', False):
        BASE_DIR = os.path.dirname(sys.executable)
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.join(BASE_DIR, 'data')
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
    DATABASE_PATH = os.path.join(DATA_DIR, 'class_points.db')
    NGROK_BIN_DIR = os.path.join(DATA_DIR, 'ngrok_bin')

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY
CORS(app)

os.makedirs(Config.DATA_DIR, exist_ok=True)
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
os.makedirs(Config.NGROK_BIN_DIR, exist_ok=True)
conf.get_default().ngrok_path = os.path.join(Config.NGROK_BIN_DIR, "ngrok.exe")

# --- 2. 数据库初始化 (单班级闭环架构) ---
def init_db():
    conn = sqlite3.connect(Config.DATABASE_PATH)
    c = conn.cursor()
    # 系统配置
    c.execute('CREATE TABLE IF NOT EXISTS system_config (id INTEGER PRIMARY KEY, class_name TEXT, teacher_name TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    # 基础业务表 (自动指向 class_id=1)
    c.execute('CREATE TABLE IF NOT EXISTS classes (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, teacher TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS groups (id INTEGER PRIMARY KEY AUTOINCREMENT, class_id INTEGER DEFAULT 1, name TEXT NOT NULL UNIQUE, color TEXT DEFAULT "#667eea")')
    c.execute('CREATE TABLE IF NOT EXISTS students (id INTEGER PRIMARY KEY AUTOINCREMENT, class_id INTEGER DEFAULT 1, group_id INTEGER, name TEXT NOT NULL, student_id TEXT NOT NULL UNIQUE, points INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS points_history (id INTEGER PRIMARY KEY AUTOINCREMENT, student_id INTEGER, change_amount INTEGER NOT NULL, reason TEXT, teacher TEXT, status TEXT DEFAULT "pending", reward_id INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS group_points_history (id INTEGER PRIMARY KEY AUTOINCREMENT, group_id INTEGER, change_amount INTEGER NOT NULL, reason TEXT, teacher TEXT, status TEXT DEFAULT "pending", created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS point_standards (id INTEGER PRIMARY KEY AUTOINCREMENT, area TEXT, category TEXT, name TEXT, default_points INTEGER, UNIQUE(area, category, name))')
    c.execute('CREATE TABLE IF NOT EXISTS rewards (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, description TEXT, points_cost INTEGER NOT NULL, image_path TEXT, stock INTEGER DEFAULT 10, is_special INTEGER DEFAULT 0, is_group_reward INTEGER DEFAULT 0, is_grocery INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS redemptions (id INTEGER PRIMARY KEY AUTOINCREMENT, student_id INTEGER, reward_id INTEGER, redeemed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS group_redemptions (id INTEGER PRIMARY KEY AUTOINCREMENT, group_id INTEGER, reward_id INTEGER, redeemed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS auctions (id INTEGER PRIMARY KEY AUTOINCREMENT, reward_id INTEGER, class_id INTEGER DEFAULT 1, status TEXT DEFAULT "active", current_price INTEGER DEFAULT 0, highest_bidder_id INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, finished_at TIMESTAMP)')
    c.execute('CREATE TABLE IF NOT EXISTS bounties (id INTEGER PRIMARY KEY AUTOINCREMENT, reward_id INTEGER, class_id INTEGER DEFAULT 1, target_points INTEGER, allowed_reasons TEXT, start_date DATE, end_date DATE, status TEXT DEFAULT "active", winner_id INTEGER, description TEXT, type TEXT DEFAULT "individual", created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, finished_at TIMESTAMP)')
    
    # --- 性能优化：添加索引 ---
    # 为积分历史表添加联合索引，加速查询和统计
    c.execute('CREATE INDEX IF NOT EXISTS idx_ph_student_status ON points_history(student_id, status)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ph_created ON points_history(created_at)')
    # 为学生表添加索引
    c.execute('CREATE INDEX IF NOT EXISTS idx_stu_group ON students(group_id)')
    
    conn.commit()
    conn.close()

def get_db_connection():
    if not os.path.exists(Config.DATABASE_PATH): init_db()
    conn = sqlite3.connect(Config.DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# --- 3. 内网穿透 (Ngrok 集成) ---
current_online_url = None
tunnel_logs = []

def log_callback(log):
    global tunnel_logs
    line = str(log).strip()
    if "t=" in line:
        msg_match = re.search(r'msg="([^"]+)"', line)
        if msg_match: tunnel_logs.append(msg_match.group(1))
    else: tunnel_logs.append(line)
    if len(tunnel_logs) > 20: tunnel_logs.pop(0)

@app.route('/api/tunnel/action', methods=['POST'])
def tunnel_action():
    global current_online_url, tunnel_logs
    data = request.json
    action = data.get('action')
    token = data.get('token', '').strip()
    # 统一使用 Config.DATA_DIR 保证便携性
    token_file = os.path.join(Config.DATA_DIR, 'ngrok_token.txt')
    
    if action == 'start':
        tunnel_logs = ["准备初始化 Ngrok..."]
        try:
            conf.get_default().log_event_callback = log_callback
            if token:
                ngrok.set_auth_token(token)
                with open(token_file, 'w') as f: f.write(token)
            elif os.path.exists(token_file):
                with open(token_file, 'r') as f: 
                    saved_token = f.read().strip()
                    ngrok.set_auth_token(saved_token)
            else: 
                return jsonify({'success': False, 'error': 'needs_token'})

            if not current_online_url:
                tunnel_logs.append("正在启动隧道进程...")
                current_online_url = ngrok.connect(5001).public_url
            return jsonify({'success': True, 'url': current_online_url})
        except Exception as e:
            error_msg = str(e)
            tunnel_logs.append(f"错误: {error_msg}")
            if "authtoken" in error_msg.lower():
                return jsonify({'success': False, 'error': 'invalid_token'})
            return jsonify({'success': False, 'error': error_msg})
    elif action == 'stop':
        try:
            ngrok.kill()
            current_online_url = None
            tunnel_logs = ["隧道已关闭"]
            return jsonify({'success': True})
        except: return jsonify({'success': False})

@app.route('/api/tunnel/status')
def tunnel_status(): return jsonify({'active': current_online_url is not None, 'url': current_online_url})

@app.route('/api/tunnel/logs')
def get_tunnel_logs(): return jsonify(tunnel_logs)

@app.route('/api/system/reset', methods=['POST'])
def system_reset():
    """彻底卸载系统：清空所有业务数据"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        tables = ['system_config', 'classes', 'groups', 'students', 'points_history', 
                  'group_points_history', 'rewards', 'redemptions', 'group_redemptions', 
                  'student_evaluations', 'auctions', 'bounties']
        for table in tables:
            cursor.execute(f'DELETE FROM {table}')
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- 4. 核心业务接口 (单班级简化版) ---

@app.route('/api/system/info')
def get_system_info():
    conn = get_db_connection()
    info = conn.execute('SELECT * FROM system_config LIMIT 1').fetchone()
    conn.close()
    return jsonify(dict(info) if info else None)

@app.route('/api/system/setup', methods=['POST'])
def system_setup():
    data = request.json
    conn = get_db_connection()
    conn.execute('INSERT OR REPLACE INTO system_config (id, class_name, teacher_name) VALUES (1, ?, ?)', (data['class_name'], data.get('teacher_name', '')))
    conn.execute('INSERT OR REPLACE INTO classes (id, name, teacher) VALUES (1, ?, ?)', (data['class_name'], data.get('teacher_name', '')))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/classes', methods=['GET'])
def get_classes():
    """获取班级 (单班级模式：始终返回 ID 为 1 的班级)"""
    try:
        conn = get_db_connection()
        c = conn.execute('SELECT * FROM classes LIMIT 1').fetchone()
        conn.close()
        return jsonify([dict(c)] if c else [])
    except Exception as e:
        return jsonify([])

@app.route('/api/students', methods=['GET', 'POST'])
def handle_students():
    conn = get_db_connection()
    if request.method == 'POST':
        data = request.json
        conn.execute('INSERT INTO students (class_id, name, student_id, group_id) VALUES (1, ?, ?, ?)', (data['name'], data['student_id'], data.get('group_id')))
        conn.commit()
        return jsonify({'success': True})
    rows = conn.execute('SELECT s.*, g.name as group_name FROM students s LEFT JOIN groups g ON s.group_id = g.id ORDER BY s.name').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/students/<int:sid>/quick_points', methods=['POST'])
def quick_points(sid):
    """直接增减积分 (跳过审核)"""
    try:
        data = request.json
        change = int(data.get('change_amount', 1))
        reason = data.get('reason', '[互动管理/随机点名] 幸运抽中加分')
        conn = get_db_connection()
        conn.execute('UPDATE students SET points = points + ? WHERE id = ?', (change, sid))
        conn.execute('INSERT INTO points_history (student_id, change_amount, reason, teacher, status) VALUES (?, ?, ?, ?, "approved")',
                     (sid, change, reason, data.get('teacher', '系统')))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups/<int:gid>/quick_points', methods=['POST'])
def group_quick_points(gid):
    """小组全员加分"""
    try:
        data = request.json
        change = int(data.get('change_amount', 1))
        reason = data.get('reason', '[互动管理/随机点名] 小组幸运抽中')
        conn = get_db_connection()
        # 1. 找到所有组员
        members = conn.execute('SELECT id FROM students WHERE group_id = ?', (gid,)).fetchall()
        for m in members:
            conn.execute('UPDATE students SET points = points + ? WHERE id = ?', (change, m['id']))
            conn.execute('INSERT INTO points_history (student_id, change_amount, reason, teacher, status) VALUES (?, ?, ?, ?, "approved")',
                         (m['id'], change, reason, data.get('teacher', '系统')))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'count': len(members)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/groups', methods=['GET', 'POST'])
def handle_groups():
    conn = get_db_connection()
    if request.method == 'POST':
        data = request.json
        conn.execute('INSERT INTO groups (class_id, name, color) VALUES (1, ?, ?)', (data['name'], data.get('color', '#667eea')))
        conn.commit()
        return jsonify({'success': True})
    rows = conn.execute('SELECT g.*, COUNT(s.id) as student_count, AVG(s.points) as avg_points FROM groups g LEFT JOIN students s ON g.id = s.group_id GROUP BY g.id').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/point_standards', methods=['GET', 'POST'])
def handle_standards():
    conn = get_db_connection()
    if request.method == 'POST':
        data = request.json
        conn.execute('INSERT INTO point_standards (area, category, name, default_points) VALUES (?, ?, ?, ?)', 
                     (data['area'], data['category'], data['name'], data['points']))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    # GET 支持按 Area 模糊匹配
    area = request.args.get('area')
    if area:
        rows = conn.execute('SELECT * FROM point_standards WHERE area LIKE ?', (f"%{area.replace('管理','')}%",)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM point_standards ORDER BY area, category').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/point_standards/<int:sid>', methods=['PUT', 'DELETE'])
def update_delete_standard(sid):
    conn = get_db_connection()
    if request.method == 'DELETE':
        conn.execute('DELETE FROM point_standards WHERE id = ?', (sid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    data = request.json
    conn.execute('UPDATE point_standards SET area=?, category=?, name=?, default_points=? WHERE id=?',
                 (data['area'], data['category'], data['name'], data['points'], sid))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/point_standards/batch_update_category', methods=['POST'])
def batch_update_std_category():
    data = request.json
    conn = get_db_connection()
    conn.execute('UPDATE point_standards SET category = ? WHERE category = ? AND area = ?',
                 (data['new_category'], data['old_category'], data['area']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/point_standards/reset', methods=['POST'])
def reset_standards():
    """重置积分理由库为最新设计的逻辑 (学科扣分 + 荣誉加分)"""
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM point_standards')
        
        subjects = ['语文', '数学', '英语', '物理', '化学', '生物', '政治', '历史', '地理']
        standards = []

        # 1. 学业管理 (学科扣分)
        for sub in subjects:
            standards.append(('学业管理', sub, '作业缺交/抄袭/敷衍', -10))
            standards.append(('学业管理', sub, '随堂测验/单元考不及格', -10))
            standards.append(('学业管理', sub, '课堂违纪(手机/睡觉/闲聊)', -5))
            standards.append(('学业管理', sub, '笔记缺失/书本未带', -2))

        # 2. 班级管理 (纪律扣分)
        attend_cats = ['早自习', '午休纪律', '晚自习', '课堂考勤', '集体活动']
        for cat in attend_cats:
            standards.append(('班级管理', cat, '迟到/早退', -5))
            standards.append(('班级管理', cat, '旷课/擅自脱岗', -20))
            standards.append(('班级管理', cat, '大声喧哗/打闹违纪', -10))

        # 3. 活动管理 (奖励项)
        act_cats = ['校级竞赛', '体育运动', '艺术文化', '社会实践']
        for cat in act_cats:
            standards.append(('活动管理', cat, '代表班级参赛(基础奖)', 5))
            standards.append(('活动管理', cat, '获得校级名次/奖项', 15))
            standards.append(('活动管理', cat, '市级及以上重大荣誉', 50))

        # 4. 自定义 (原德育/其他加分项)
        plus_cats = ['品德楷模', '班级勤务', '同伴互助']
        for cat in plus_cats:
            standards.append(('自定义', cat, '拾金不昧/见义勇为', 20))
            standards.append(('自定义', cat, '主动承担额外扫除', 5))
            standards.append(('自定义', cat, '辅导同学学业(获认可)', 10))
            
        # 5. 自定义 (负分项)
        standards.append(('自定义', '行为规范', '损坏公物/破坏环境', -10))
        standards.append(('自定义', '行为规范', '浪费粮食/水电', -5))
        standards.append(('自定义', '行为规范', '仪容仪表不整', -2))

        conn.executemany('INSERT INTO point_standards (area, category, name, default_points) VALUES (?, ?, ?, ?)', standards)
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/point_standards/export', methods=['GET'])
def export_standards():
    """导出积分理由库到 Excel"""
    try:
        conn = get_db_connection()
        rows = conn.execute('SELECT area, category, name, default_points FROM point_standards ORDER BY area, category').fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "积分评分标准"
        ws.append(["业务大类", "项目分类", "事项名称", "默认分值"])
        
        for r in rows:
            ws.append([r['area'], r['category'], r['name'], r['default_points']])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="积分评分标准.xlsx")
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/point_standards/import', methods=['POST'])
def import_standards():
    """从 Excel 导入积分理由库"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400
    
    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'error': '无效文件'}), 400

    try:
        wb = load_workbook(file)
        ws = wb.active
        
        standards = []
        # 跳过表头，从第二行开始
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[2]: continue # 跳过空行
            standards.append((row[0], row[1], row[2], int(row[3] or 0)))

        if not standards:
            return jsonify({'error': '文件中没有有效数据'}), 400

        conn = get_db_connection()
        # 导入通常是增量还是覆盖？这里采用覆盖逻辑，保持与 reset 一致
        conn.execute('DELETE FROM point_standards')
        conn.executemany('INSERT INTO point_standards (area, category, name, default_points) VALUES (?, ?, ?, ?)', standards)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'count': len(standards)})
    except Exception as e:
        return jsonify({'error': f"解析文件失败: {str(e)}"}), 500

@app.route('/api/rewards', methods=['GET', 'POST'])
def handle_rewards():
    """奖品管理：获取、添加 (支持图片上传)"""
    conn = get_db_connection()
    if request.method == 'POST':
        # 兼容 JSON 和 Form-Data
        if request.is_json:
            data = request.json
            name, desc, pts, stock, is_g, is_s = data['name'], data.get('description',''), data['points_cost'], data.get('stock', 10), data.get('is_grocery', 0), data.get('is_special', 0)
            img_path = ''
        else:
            name = request.form.get('name')
            desc = request.form.get('description', '')
            pts = request.form.get('points_cost', 0)
            stock = request.form.get('stock', 10)
            is_g = request.form.get('is_grocery', 0)
            is_s = request.form.get('is_special', 0)
            img_path = ''
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename:
                    fname = secure_filename(f"{int(time.time())}_{file.filename}")
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
                    img_path = f"/static/uploads/{fname}"

        conn.execute('INSERT INTO rewards (name, description, points_cost, stock, is_grocery, is_special, image_path, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                     (name, desc, int(pts), int(stock), int(is_g), int(is_s), img_path, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    # GET: 支持小铺筛选
    is_grocery = request.args.get('is_grocery')
    if is_grocery is not None:
        try:
            is_grocery = int(is_grocery)
        except:
            is_grocery = 0
        rows = conn.execute('SELECT * FROM rewards WHERE is_grocery = ? ORDER BY created_at DESC', (is_grocery,)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM rewards ORDER BY created_at DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/rewards/<int:rid>', methods=['DELETE'])
def delete_reward(rid):
    conn = get_db_connection()
    conn.execute('DELETE FROM rewards WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/classes/<int:class_id>/stats', methods=['GET'])
def get_class_stats(class_id):
    """获取班级统计信息 (支持日期筛选，区分正负分，荣誉榜聚合)"""
    try:
        date_str = request.args.get('date')
        conn = get_db_connection()
        
        # 1. 基础汇总
        res_stats = conn.execute('SELECT AVG(points) as avg, MAX(points) as max, MIN(points) as min, COUNT(*) as count FROM students').fetchone()
        
        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # 2. 获取加分记录 (荣誉榜)
        all_plus = conn.execute('''
            SELECT ph.*, s.name as student_name 
            FROM points_history ph JOIN students s ON ph.student_id = s.id 
            WHERE ph.status = 'approved' AND ph.change_amount > 0 AND date(ph.created_at) = ?
            AND ph.reason NOT LIKE '%兑换%' 
            AND ph.reason NOT LIKE '%拍卖%' 
            AND ph.reason NOT LIKE '%悬赏%'
            AND ph.reason NOT LIKE '%结项%'
            ORDER BY ph.created_at DESC
        ''', (date_str,)).fetchall()
        
        # --- 荣誉榜聚合逻辑 ---
        plus_list = []
        benchmark_groups = {} # key: (reason, created_at) -> {count, reason, time, amount}
        
        for p in all_plus:
            p_dict = dict(p)
            if p_dict['reason'].startswith('[基本准则]'):
                key = (p_dict['reason'], p_dict['created_at'])
                if key not in benchmark_groups:
                    benchmark_groups[key] = {
                        'student_name': '全班达标', # 初始占位
                        'reason': p_dict['reason'],
                        'change_amount': p_dict['change_amount'],
                        'created_at': p_dict['created_at'],
                        'count': 0
                    }
                benchmark_groups[key]['count'] += 1
            else:
                plus_list.append(p_dict)
        
        # 将聚合后的基本准则加入列表
        for g in benchmark_groups.values():
            g['student_name'] = f"达标 {g['count']} 人"
            plus_list.append(g)
        
        # 重新按时间排序
        plus_list.sort(key=lambda x: x['created_at'], reverse=True)

        # 3. 获取减分记录 (违纪榜 - 保持明细展示)
        minus = conn.execute('''
            SELECT ph.*, s.name as student_name 
            FROM points_history ph JOIN students s ON ph.student_id = s.id 
            WHERE ph.status = 'approved' AND ph.change_amount < 0 AND date(ph.created_at) = ?
            AND ph.reason NOT LIKE '%兑换%' 
            AND ph.reason NOT LIKE '%拍卖%' 
            AND ph.reason NOT LIKE '%悬赏%'
            AND ph.reason NOT LIKE '%结项%'
            ORDER BY ph.created_at DESC
        ''', (date_str,)).fetchall()
        
        class_info = conn.execute('SELECT * FROM system_config LIMIT 1').fetchone()
        conn.close()
        
        return jsonify({
            'class_info': dict(class_info) if class_info else {'name': '未设置'},
            'total_students': res_stats['count'],
            'avg_points': round(res_stats['avg'] or 0, 1),
            'max_points': res_stats['max'] or 0,
            'min_points': res_stats['min'] or 0,
            'plus_changes': plus_list,
            'minus_changes': [dict(r) for r in minus]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/system/stats', methods=['GET'])
def get_system_stats():
    return get_class_stats(1)

@app.route('/api/ranking', methods=['GET'])
def get_ranking_api():
    """获取单班级排行榜 (优化版 SQL)"""
    try:
        rtype = request.args.get('type', 'student')
        start = request.args.get('start_date')
        end = request.args.get('end_date')
        
        conn = get_db_connection()
        date_filter = ""
        params = []
        if start and end:
            date_filter = " AND date(ph.created_at) BETWEEN ? AND ?"
            params = [start, end]

        if rtype == 'student':
            if date_filter:
                sql = f'''
                    SELECT s.id, s.name, s.student_id, g.name as group_name,
                           COALESCE(SUM(ph.change_amount), 0) as points
                    FROM students s
                    LEFT JOIN groups g ON s.group_id = g.id
                    LEFT JOIN points_history ph ON s.id = ph.student_id AND ph.status = 'approved' {date_filter}
                    GROUP BY s.id ORDER BY points DESC, s.name ASC
                '''
            else:
                sql = '''
                    SELECT id, name, student_id, points, 
                           (SELECT name FROM groups WHERE id = students.group_id) as group_name
                    FROM students 
                    ORDER BY points DESC, name ASC
                '''
        else: # 小组榜
            if date_filter:
                sql = f'''
                    SELECT g.id, g.name, g.color,
                           COALESCE(SUM(ph.change_amount), 0) as points
                    FROM groups g
                    LEFT JOIN students s ON g.id = s.group_id
                    LEFT JOIN points_history ph ON s.id = ph.student_id AND ph.status = 'approved' {date_filter}
                    GROUP BY g.id ORDER BY points DESC, g.name ASC
                '''
            else:
                sql = '''
                    SELECT id, name, color,
                           (SELECT SUM(points) FROM students WHERE group_id = groups.id) as points
                    FROM groups 
                    ORDER BY points DESC, name ASC
                '''
        
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        import traceback
        traceback.print_exc() # 打印报错到黑窗口
        return jsonify({'error': str(e)}), 500

# ============ 互动管理 API (拍卖、悬赏) ============

@app.route('/api/auction/start', methods=['POST'])
def start_auction():
    data = request.json
    conn = get_db_connection()
    # 取消旧拍卖，开启新拍卖
    conn.execute('UPDATE auctions SET status = "cancelled" WHERE status = "active"')
    conn.execute('INSERT INTO auctions (reward_id, class_id, current_price, status) VALUES (?, 1, ?, "active")',
                 (data['reward_id'], data.get('start_price', 0)))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/auction/current', methods=['GET'])
def get_current_auction():
    try:
        conn = get_db_connection()
        # 关联奖励表、学生表、班级配置
        row = conn.execute('''
            SELECT a.*, r.name as reward_name, r.image_path, r.description,
                   s.name as bidder_name,
                   (SELECT class_name FROM system_config LIMIT 1) as class_name
            FROM auctions a 
            JOIN rewards r ON a.reward_id = r.id
            LEFT JOIN students s ON a.highest_bidder_id = s.id
            WHERE a.status = "active" 
            ORDER BY a.created_at DESC LIMIT 1
        ''').fetchone()
        conn.close()
        return jsonify(dict(row) if row else None)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auction/bid', methods=['POST'])
def place_bid():
    data = request.json
    conn = get_db_connection()
    auc = conn.execute('SELECT * FROM auctions WHERE id = ? AND status = "active"', (data['auction_id'],)).fetchone()
    if not auc or data['amount'] <= auc['current_price']:
        return jsonify({'error': '竞价已失效或出价过低'}), 400
    
    conn.execute('UPDATE auctions SET current_price = ?, highest_bidder_id = ? WHERE id = ?',
                 (data['amount'], data['student_id'], data['auction_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/auction/finish', methods=['POST'])
def finish_auction():
    data = request.json
    conn = get_db_connection()
    auc = conn.execute('SELECT a.*, r.name as rname FROM auctions a JOIN rewards r ON a.reward_id = r.id WHERE a.id = ?', (data['auction_id'],)).fetchone()
    if auc and auc['highest_bidder_id']:
        # 扣除积分
        conn.execute('UPDATE students SET points = points - ? WHERE id = ?', (auc['current_price'], auc['highest_bidder_id']))
        # 记录历史
        conn.execute('INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, "拍卖系统")',
                     (auc['highest_bidder_id'], -auc['current_price'], f"拍卖得标: {auc['rname']}"))
    
    conn.execute('UPDATE auctions SET status = "finished", finished_at = ? WHERE id = ?',
                 (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), data['auction_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/bounty/start', methods=['POST'])
def start_bounty():
    """开启悬赏 (补全字段)"""
    try:
        data = request.json
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO bounties (
                reward_id, class_id, target_points, type, description, 
                allowed_reasons, start_date, end_date, status
            ) VALUES (?, 1, ?, ?, ?, ?, ?, ?, "active")
        ''', (
            data['reward_id'], data['target_points'], data.get('type', 'individual'),
            data.get('description', ''), data.get('allowed_reasons', ''),
            data.get('start_date'), data.get('end_date')
        ))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bounties/progress')
def get_bounties_progress():
    """获取悬赏进度 (精准规则匹配版)"""
    try:
        conn = get_db_connection()
        today = datetime.now().strftime('%Y-%m-%d')
        # 1. 查找活跃悬赏
        rows = conn.execute('''
            SELECT b.*, r.name as reward_name, r.points_cost as reward_prize, r.stock 
            FROM bounties b 
            JOIN rewards r ON b.reward_id = r.id 
            WHERE b.status = "active" AND (b.end_date IS NULL OR date(b.end_date) >= ?)
        ''', (today,)).fetchall()
        
        res = []
        for b in rows:
            # 理由过滤条件
            reasons = b['allowed_reasons'].split(',') if b['allowed_reasons'] else []
            reason_filter = ""
            params = []
            if reasons:
                placeholders = ','.join(['?'] * len(reasons))
                reason_filter = f" AND reason IN ({placeholders})"
                params = reasons
            
            if b['type'] == 'group':
                # 小组：在该悬赏规则下的累计加分 (取前三)
                sql = f'''
                    SELECT g.id, g.name, SUM(ph.change_amount) as current_points
                    FROM groups g
                    JOIN students s ON g.id = s.group_id
                    JOIN points_history ph ON s.id = ph.student_id
                    WHERE ph.status = 'approved' AND ph.change_amount > 0 {reason_filter}
                    GROUP BY g.id ORDER BY current_points DESC LIMIT 3
                '''
                leader_rows = conn.execute(sql, params).fetchall()
            else:
                # 个人：在该悬赏规则下的累计加分 (取前三)
                sql = f'''
                    SELECT s.id, s.name, SUM(ph.change_amount) as current_points
                    FROM students s
                    JOIN points_history ph ON s.id = ph.student_id
                    WHERE ph.status = 'approved' AND ph.change_amount > 0 {reason_filter}
                    GROUP BY s.id ORDER BY current_points DESC LIMIT 3
                '''
                leader_rows = conn.execute(sql, params).fetchall()

            res.append({
                'id': b['id'],
                'reward_name': b['reward_name'],
                'reward_prize': b['reward_prize'],
                'target_points': b['target_points'],
                'type': b['type'],
                'stock': b['stock'],
                'leaders': [
                    {'id': r['id'], 'name': r['name'], 'points': r['current_points']} 
                    for r in leader_rows
                ]
            })
        conn.close()
        return jsonify(res)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bounty/preview_finish', methods=['POST'])
def preview_bounty_finish():
    """预览结项扣分方案"""
    data = request.json
    bid = data.get('bounty_id')
    leader_id = data.get('leader_id')
    try:
        conn = get_db_connection()
        b = conn.execute('SELECT b.*, r.points_cost, r.name as rname FROM bounties b JOIN rewards r ON b.reward_id = r.id WHERE b.id = ?', (bid,)).fetchone()
        if not b: return jsonify({'error': '悬赏不存在'}), 404

        plan = []
        total_pts = b['points_cost']
        if b['type'] == 'group':
            members = conn.execute('SELECT id, name, points FROM students WHERE group_id = ? ORDER BY points DESC, name ASC', (leader_id,)).fetchall()
            if not members: return jsonify({'error': '该小组无成员'}), 400
            
            n = len(members)
            base = total_pts // n
            extra = total_pts % n
            
            for i, m in enumerate(members):
                deduct = base + (1 if i < extra else 0)
                plan.append({'student_id': m['id'], 'name': m['name'], 'current': m['points'], 'deduct': deduct})
        else:
            m = conn.execute('SELECT id, name, points FROM students WHERE id = ?', (leader_id,)).fetchone()
            plan.append({'student_id': m['id'], 'name': m['name'], 'current': m['points'], 'deduct': total_pts})
            
        conn.close()
        return jsonify({'reward_name': b['rname'], 'plan': plan, 'total': total_pts})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bounty/finish', methods=['POST'])
def finish_bounty():
    """执行结项：按方案扣分"""
    data = request.json
    bid = data.get('bounty_id')
    plan = data.get('plan', []) # 前端确认后的扣分方案
    
    try:
        conn = get_db_connection()
        b = conn.execute('SELECT * FROM bounties WHERE id = ?', (bid,)).fetchone()
        
        # 1. 严格按方案执行扣分
        for item in plan:
            conn.execute('UPDATE students SET points = points - ? WHERE id = ?', (item['deduct'], item['student_id']))
            conn.execute('INSERT INTO points_history (student_id, change_amount, reason, teacher) VALUES (?, ?, ?, "悬赏结项")',
                         (item['student_id'], -item['deduct'], f"达成悬赏: {data.get('reward_name')}"))

        # 2. 扣除奖品库存
        conn.execute('UPDATE rewards SET stock = stock - 1 WHERE id = ?', (b['reward_id'],))
        
        # 3. 标记悬赏状态
        conn.execute('UPDATE bounties SET status = "finished", winner_id = ?, finished_at = ? WHERE id = ?',
                     (data.get('leader_id'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), bid))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/events/recent')
def get_events_recent():
    """获取最近荣誉动态 (支持日期筛选)"""
    try:
        date_str = request.args.get('date')
        conn = get_db_connection()
        
        date_filter = ""
        params = []
        if date_str:
            date_filter = " AND date(ph.created_at) = ?"
            params = [date_str]

        # 筛选逻辑：负分变动 且 包含特定关键词
        rows = conn.execute(f'''
            SELECT ph.reason as reward_name, s.name as winner_name, 
                   ph.created_at as time,
                   CASE 
                     WHEN ph.reason LIKE '拍卖%' THEN 'auction'
                     WHEN ph.reason LIKE '达成悬赏%' THEN 'bounty'
                     ELSE 'reward'
                   END as type
            FROM points_history ph
            JOIN students s ON ph.student_id = s.id
            WHERE ph.change_amount < 0 
            AND (ph.reason LIKE '拍卖%' OR ph.reason LIKE '达成悬赏%' OR ph.reason LIKE '兑换%')
            {date_filter}
            ORDER BY ph.created_at DESC LIMIT 20
        ''', params).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- 5. 积分与审核 ---

@app.route('/api/audit/submit', methods=['POST'])
def submit_audit():
    try:
        data = request.json
        conn = get_db_connection()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        change_amount = int(data.get('change_amount', 0))
        submitted_ids = set(data.get('student_ids', []))
        reason = data.get('reason', '自助申报')
        submitter = data.get('submitter', '自助')
        
        # 核心判断：是否触发“基本准则”逻辑 (负分 且 属于学业/班级管理)
        is_benchmark_rule = change_amount < 0 and ('[学业管理' in reason or '[班级管理' in reason)

        if is_benchmark_rule:
            # === 模式 A：基本准则管理 (自动生效，无需审核) ===
            # 1. 扣分部分：仅针对名单内的人
            if submitted_ids:
                neg_records = []
                for sid in submitted_ids:
                    # 状态直接为 approved
                    neg_records.append((sid, change_amount, reason, submitter, "approved", now))
                    # 实时扣分
                    conn.execute('UPDATE students SET points = points + ? WHERE id = ?', (change_amount, sid))
                conn.executemany('INSERT INTO points_history (student_id, change_amount, reason, teacher, status, created_at) VALUES (?, ?, ?, ?, ?, ?)', neg_records)

            # 2. 奖励部分：全班 - 扣分名单 = 达标名单
            all_students = conn.execute('SELECT id FROM students').fetchall()
            all_ids = {s['id'] for s in all_students}
            bonus_ids = all_ids - submitted_ids
            
            if bonus_ids:
                try:
                    simple_reason = reason.split('] ')[-1]
                except:
                    simple_reason = "日常规范"
                bonus_reason = f"[基本准则] {simple_reason} - 达标奖励"
                
                bonus_records = []
                for bid in bonus_ids:
                    # 状态直接为 approved
                    bonus_records.append((bid, 2, bonus_reason, f"系统({submitter})", "approved", now))
                    # 实时加分
                    conn.execute('UPDATE students SET points = points + 2 WHERE id = ?', (bid,))
                conn.executemany('INSERT INTO points_history (student_id, change_amount, reason, teacher, status, created_at) VALUES (?, ?, ?, ?, ?, ?)', bonus_records)

        else:
            # === 模式 B：普通加减分 (荣誉/自定义等) ===
            # 仅针对选中的人，不触发全员逻辑
            if not submitted_ids:
                return jsonify({'error': '未选择学生'}), 400
                
            records = []
            for sid in submitted_ids:
                records.append((sid, change_amount, reason, submitter, "pending", now))
            conn.executemany('INSERT INTO points_history (student_id, change_amount, reason, teacher, status, created_at) VALUES (?, ?, ?, ?, ?, ?)', records)

        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/audit/pending')
def get_pending():
    conn = get_db_connection()
    rows = conn.execute('SELECT ph.*, s.name as student_name FROM points_history ph JOIN students s ON ph.student_id = s.id WHERE ph.status = "pending"').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/audit/process', methods=['POST'])
def process_audit():
    data = request.json
    conn = get_db_connection()
    for aid in data['audit_ids']:
        if data['action'] == 'approve':
            record = conn.execute('SELECT * FROM points_history WHERE id = ?', (aid,)).fetchone()
            conn.execute('UPDATE students SET points = points + ? WHERE id = ?', (record['change_amount'], record['student_id']))
            conn.execute('UPDATE points_history SET status = "approved" WHERE id = ?', (aid,))
        else:
            conn.execute('UPDATE points_history SET status = "rejected" WHERE id = ?', (aid,))
    conn.commit()
    return jsonify({'success': True})

# --- 5. 权限与路由 ---

@app.before_request
def check_auth():
    # 终极简化版白名单 (加入排行榜、学生、小组、申报、彩蛋等接口)
    allowed = ['/login', '/static', '/student_portal', '/grocery_shop', '/auction', '/bounties', '/author',
               '/api/system/info', '/api/system/setup', '/api/students', '/api/groups', 
               '/api/point_standards', '/api/audit/submit', '/api/rewards', '/api/tunnel',
               '/api/auction/current', '/api/bounties/progress', '/api/events/recent', '/api/ranking']
    if any(request.path.startswith(p) for p in allowed): return
    if 'logged_in' not in session: return redirect(url_for('login'))

@app.route('/')
def index(): return render_template('index.html')

@app.route('/student_portal')
def student_portal(): return render_template('student_portal.html')

@app.route('/grocery_shop')
def grocery_shop(): return render_template('grocery_shop.html')

@app.route('/author')
def egg_page(): return render_template('egg.html')

# --- 管理端页面路由 ---
@app.route('/students')
def students_page(): return render_template('students.html')

@app.route('/points')
def points_page(): return render_template('points.html')

@app.route('/ranking')
def ranking_page(): return render_template('ranking.html')

@app.route('/auction')
def auction_page(): return render_template('auction.html')

@app.route('/random')
def random_page(): return render_template('random.html')

@app.route('/favicon.ico')
def favicon(): return send_from_directory(os.path.join(app.root_path, 'static'), 'favicon.ico', mimetype='image/vnd.microsoft.icon')

def get_system_password():
    """从本地文件读取系统密码"""
    pwd_file = os.path.join(Config.BASE_DIR, 'password.txt')
    if not os.path.exists(pwd_file):
        with open(pwd_file, 'w', encoding='utf-8') as f: f.write('123456')
        return '123456'
    with open(pwd_file, 'r', encoding='utf-8') as f: return f.read().strip()

@app.route('/api/students/<int:sid>/history', methods=['GET'])
def get_student_history(sid):
    """获取单个学生的积分明细"""
    try:
        conn = get_db_connection()
        # 1. 获取基本信息和排名
        all_students = conn.execute('SELECT id, points FROM students ORDER BY points DESC, name ASC').fetchall()
        rank = -1
        points = 0
        for i, s in enumerate(all_students):
            if s['id'] == sid:
                rank = i + 1
                points = s['points']
                break
        
        # 2. 获取最近 20 条历史
        history = conn.execute('''
            SELECT ph.* FROM points_history ph 
            WHERE ph.student_id = ? AND ph.status = 'approved'
            ORDER BY ph.created_at DESC LIMIT 20
        ''', (sid,)).fetchall()
        
        conn.close()
        return jsonify({
            'points': points,
            'rank': rank,
            'history': [dict(h) for h in history]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/verify_password', methods=['POST'])
def verify_password_api():
    data = request.json
    if str(data.get('password')) == get_system_password():
        return jsonify({'success': True})
    return jsonify({'success': False}), 403

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if str(request.json.get('password')) == get_system_password():
            session['logged_in'] = True
            return jsonify({'success': True})
        return jsonify({'success': False}), 403
    return render_template('login.html')

@app.route('/api/students/template', methods=['GET'])
def download_student_template():
    """下载学生导入模板 (带细则与样例)"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "学生名单填报"
        
        # 1. 写入表头
        headers = ["姓名", "学号", "分组", "初始积分"]
        ws.append(headers)
        
        # 2. 写入样例数据
        sample = ["张小明", "2025001", "飞龙组", 100]
        ws.append(sample)
        
        # 3. 增加说明工作表
        ws_info = wb.create_sheet("导入细则(必读)")
        instructions = [
            ["项目", "要求", "示例"],
            ["姓名", "学生真实姓名，必填", "张小明"],
            ["学号", "系统唯一识别码，不可重复，必填", "2025001"],
            ["分组", "所属小组名称，选填。若填写的小组不存在，系统会自动为您创建。", "飞龙组"],
            ["初始积分", "学生初始的分数，选填，默认为 0", "100"]
        ]
        for row in instructions:
            ws_info.append(row)
            
        # 设置列宽美化
        for sheet in [ws, ws_info]:
            for col in ['A', 'B', 'C', 'D']:
                sheet.column_dimensions[col].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="班级成员导入模板.xlsx")
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':

    init_db()

    

    # 自动获取局域网 IP (方便手机访问)

    try:

        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)

        s.connect(('8.8.8.8', 80))

        local_ip = s.getsockname()[0]

        s.close()

    except: local_ip = '127.0.0.1'



    port = 5001

    url = f"http://localhost:{port}"

    print("=" * 60)

    print("班级积分管理系统 V3 - 单班级专业版")

    print(f"【电脑端访问】: {url}")

    print(f"【手机端访问】: http://{local_ip}:{port}")

    print("=" * 60)



    # 仅在非重载模式下自动打开浏览器

    if not os.environ.get("WERKZEUG_RUN_MAIN"):

        try: webbrowser.open_new(url)

        except: pass



    app.run(host='0.0.0.0', port=port, debug=False)
